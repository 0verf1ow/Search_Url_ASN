# _*_ coding:utf-8 _*_
from openpyxl  import Workbook
import requests
import re
import socket
import json
import time
import sys
import argparse


"""域名解析成ip地址"""
def domain2ip(domain):
    try:
        ip = socket.gethostbyname(domain)
        return ip
    except:
        return None

"""传入ip地址查询归属"""
def getASN(ip):
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.89 Safari/537.36'}
    url = 'https://www.ip138.com/iplookup.asp?ip={}&action=2'.format(ip)
    try:
        r = requests.get(url, headers = headers)
        r.raise_for_status()
        r.encoding = "gb2312"
        str_data = re.search(r'var ip_result = (.*)', r.text).group(1)[:-2]
        json_data = json.loads(str_data)
        return json_data
    except:
        exit("查询接口异常")


"""从数据中提取出ip"""
def getIP(url):
    try:
        """匹配域名"""
        domain = re.search(r'[a-zA-Z]+://([^\s]*\.[a-zA-Z]+)', url).group(1)
        ip = domain2ip(domain)
    except:
        try:
            """匹配IP地址"""
            ip = re.search(
                r'\b(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)',
                url).group()
        except:
            ip = None
    return ip

"""写入excel文件"""
def xls(data, sava_name):
    wb = Workbook()
    ws = wb.active

    # 表格宽度
    ws.column_dimensions['A'].width = 33.333
    ws.column_dimensions['B'].width = 8.888
    ws.column_dimensions['C'].width = 33.333
    ws.column_dimensions['D'].width = 46.666
    ws.column_dimensions['E'].width = 33.333

    # 表头
    ws['A1'] = 'URL'
    ws['B1'] = '运营商'
    ws['C1'] = '归属'
    ws['D1'] = 'ASN'
    ws['E1'] = 'IP段'

    for j in data:
        ws.append([j['url'], j['ip_c_list'][0]['yunyin'], j['ip_c_list'][0]['idc'], j['ASN归属地'], j['iP段']])

    wb.save(sava_name)

"""打印 Banner"""
def print_banner():
    banner = """
   _____ _                     _____ _   _ 
  / ____| |             /\    / ____| \ | |
 | |    | |__   __ _   /  \  | (___ |  \| |
 | |    | '_ \ / _` | / /\ \  \___ \| . ` |
 | |____| | | | (_| |/ ____ \ ____) | |\  |
  \_____|_| |_|\__,_/_/    \_\_____/|_| \_|
                                           
__Author__:    Version:0.1   Runtime:{}                               
        """.format(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    print(banner)

def main(f = None, url = '', sleep_time = 0, save_name = 'save.xlsx'):
    if f:
        data = []
        for url in f.readlines():
            ip = getIP(url)
            if ip:
                ASN = getASN(ip)
                ASN['url'] = url
                data.append(ASN)
                print("[*] 正在查：{}  ==>  {}".format(url.strip(), ASN['ip_c_list'][0]['idc']))
            else:
                print("[!] 解析不到ip：" + url.strip())

            time.sleep(sleep_time)

        xls(data, save_name)
    elif url:
        ip = getIP(url)
        ASN = getASN(ip)
        print("[*] 查询结果如下：")
        print('URL：{}'.format(url.strip()))
        print('运营商：{}'.format(ASN['ip_c_list'][0]['yunyin']))
        print('归属：{}'.format(ASN['ip_c_list'][0]['idc']))
        print('ASN：{}'.format(ASN['ASN归属地']))
        print('IP段：{}'.format(ASN['iP段']))

if __name__ == '__main__':
    print_banner()
    parser = argparse.ArgumentParser()
    parser.add_argument('-u', '--url')
    parser.add_argument('-f', '--file')
    parser.add_argument('-o', '--ouput', default='save.xlsx')
    args = parser.parse_args()
    sleep_time = 3  # 查一个ip间的间隔，太快会被banip,尽量长点，单位 秒
    use = """Example: python3 ChaASN.py -u https://www.baidu.com
         python3 ChaASN.py -u http://127.0.0.1:666
         python3 ChaASn.py -f url.txt -o save.xlsx
        """
    if len(sys.argv) < 2:
        exit(use)
    elif args.url:
        main(url = args.url)
    elif args.file:
        f = open(args.file,'r')
        main(f = f,sleep_time = sleep_time, save_name = args.ouput)
        f.close()
    else:
        print(use)
